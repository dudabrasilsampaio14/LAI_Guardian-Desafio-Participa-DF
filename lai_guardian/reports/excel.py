from __future__ import annotations
"""Exportador de Excel em padrão institucional (resumo + auditoria), com formatação voltada a leitura e controle."""

import datetime
from typing import Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Paleta discreta e institucional (boa para leitura e para relatório de controle).
CGDF_BLUE = "003366"
CGDF_BLUE_DARK = "002244"
CGDF_GRAY = "F2F4F7"
CGDF_GRAY_DARK = "D0D7DE"
WHITE = "FFFFFF"

RISK_COLOR = {
    "CRÍTICO": "B00020",  # vermelho forte
    "ALTO": "D97706",     # laranja
    "MÉDIO": "CA8A04",    # âmbar
    "BAIXO": "15803D",    # verde
    "": "6B7280",
    None: "6B7280",
}

THIN = Side(style="thin", color=CGDF_GRAY_DARK)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor=CGDF_BLUE)
    font = Font(color=WHITE, bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 26
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = BORDER


def _style_body(ws, nrows: int, ncols: int) -> None:
    body_font = Font(color="111827", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    zebra = PatternFill("solid", fgColor=CGDF_GRAY)

    for r in range(2, nrows + 1):
        is_zebra = (r % 2 == 0)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = BORDER
            cell.alignment = wrap
            if is_zebra:
                cell.fill = zebra

    # Altura para leitura (limite para evitar arquivo pesado)
    for r in range(2, min(nrows + 1, 3000)):
        ws.row_dimensions[r].height = 42


def _auto_width(ws, max_rows_scan: int = 2000) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[:max_rows_scan]:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 70)


def _add_conditional_formatting(ws, header_map: Dict[str, int], nrows: int) -> None:
    # Contém dados pessoais: verde (False) / vermelho (True)
    if "Contem_Dados_Pessoais" in header_map:
        col = get_column_letter(header_map["Contem_Dados_Pessoais"])
        rng = f"{col}2:{col}{nrows}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'=${col}2=TRUE'], fill=PatternFill("solid", fgColor="FEE2E2")),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'=${col}2=FALSE'], fill=PatternFill("solid", fgColor="DCFCE7")),
        )

    # Risco_Max: cor por criticidade
    if "Risco_Max" in header_map:
        col = get_column_letter(header_map["Risco_Max"])
        rng = f"{col}2:{col}{nrows}"
        for risk, color in RISK_COLOR.items():
            if risk in ("", None):
                continue
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f'=${col}2="{risk}"'],
                    fill=PatternFill("solid", fgColor=color),
                    font=Font(color=WHITE, bold=True),
                ),
            )


def _create_summary_sheet(wb, df: pd.DataFrame, sheet_name: str = "resumo") -> None:
    ws = wb.create_sheet(sheet_name, 0)

    title_fill = PatternFill("solid", fgColor=CGDF_BLUE_DARK)
    title_font = Font(color=WHITE, bold=True, size=14)
    subtitle_font = Font(color="111827", bold=True, size=11)

    ws["A1"] = "LAI Guardian — Resumo Executivo"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Data/Hora do Relatório"
    ws["B3"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws["A3"].font = subtitle_font

    total = int(len(df))
    positives = int(df.get("Contem_Dados_Pessoais", pd.Series([False] * total)).sum()) if total else 0
    pct = (positives / total) if total else 0.0

    ws["A5"] = "Total de Registros"
    ws["B5"] = total
    ws["A6"] = "Registros com Dados Pessoais"
    ws["B6"] = positives
    ws["A7"] = "Percentual com Dados Pessoais"
    ws["B7"] = pct
    ws["B7"].number_format = "0.00%"

    for r in range(5, 8):
        ws[f"A{r}"].font = subtitle_font
        ws[f"A{r}"].alignment = Alignment(horizontal="left")
        ws[f"B{r}"].alignment = Alignment(horizontal="left")

    ws["A9"] = "Distribuição por Risco"
    ws["A9"].font = subtitle_font

    risk_series = df.get("Risco_Max", pd.Series([""] * total)).fillna("").astype(str)
    counts = risk_series.value_counts().to_dict()

    rows = [
        ("CRÍTICO", counts.get("CRÍTICO", 0)),
        ("ALTO", counts.get("ALTO", 0)),
        ("MÉDIO", counts.get("MÉDIO", 0)),
        ("BAIXO", counts.get("BAIXO", 0)),
        ("(vazio)", counts.get("", 0)),
    ]

    ws["A10"] = "Risco"
    ws["B10"] = "Qtd"
    for cell in ("A10", "B10"):
        ws[cell].font = Font(color=WHITE, bold=True)
        ws[cell].fill = PatternFill("solid", fgColor=CGDF_BLUE)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].border = BORDER

    for i, (risk, val) in enumerate(rows, start=1):
        rr = 10 + i
        ws[f"A{rr}"] = risk
        ws[f"B{rr}"] = int(val)
        ws[f"A{rr}"].border = BORDER
        ws[f"B{rr}"].border = BORDER
        ws[f"A{rr}"].alignment = Alignment(horizontal="left")
        ws[f"B{rr}"].alignment = Alignment(horizontal="center")

        if risk in RISK_COLOR:
            ws[f"A{rr}"].fill = PatternFill("solid", fgColor=RISK_COLOR[risk])
            ws[f"A{rr}"].font = Font(color=WHITE, bold=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.freeze_panes = "A5"


def export_excel(df: pd.DataFrame, path: str) -> None:
    """
    Exporta um relatório Excel premium (banca/CGDF/TCU).

    Inclui:
      - Aba "resumo" (executivo)
      - Aba "auditoria" (linha a linha)
      - Cabeçalho institucional, zebra striping, bordas, filtros
      - Formatação condicional (Risco_Max e Contem_Dados_Pessoais)
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="auditoria")
        ws = writer.sheets["auditoria"]
        wb = writer.book

        # Resumo executivo (primeira aba)
        _create_summary_sheet(wb, df, sheet_name="resumo")

        # Mapear headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_map = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}

        # Navegação
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Estilo
        _style_header(ws, ws.max_column)
        _style_body(ws, ws.max_row, ws.max_column)
        _auto_width(ws)
        _add_conditional_formatting(ws, header_map, ws.max_row)

        # Ajustes pontuais para leitura
        for name in ("Texto_Analise", "Texto Mascarado", "Versao_Publicavel"):
            if name in header_map:
                col = get_column_letter(header_map[name])
                ws.column_dimensions[col].width = 70

        if "Qtd_Achados" in header_map:
            col = get_column_letter(header_map["Qtd_Achados"])
            for r in range(2, ws.max_row + 1):
                ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
